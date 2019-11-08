import sys
import os
from openpyxl import *
from openpyxl.styles import *
import datetime
import random
from datetime import *
from time import *
from itertools import cycle
from tournament.models import Tournament, Game
from django.utils import timezone
from django.db.models import Q
from ping.settings import BASE_DIR


def get_current_tournament():
    if Tournament.objects.exists():
        return Tournament.objects.latest('id')
    return None


def get_number_of_rounds(tour):
    if tour.number_of_participants() % 2 == 0:
        return tour.games_per_person
    else:
        return tour.games_per_person // 2


def generate_games_subset(games, participants):
    tournament = get_current_tournament()
    random.shuffle(participants)
    number_of_participants = len(participants)
    games_subset = []

    for i in range(number_of_participants // 2):

        game = Game(tournament=tournament, id1=participants[i], id2=participants[number_of_participants - i - 1])

        if game not in games:
            games_subset.append(game)

    return games_subset


def number_of_games(games, p_id):
    number = 0
    for game in games:
        if game.id1 == p_id or game.id2 == p_id:
            number += 1
    return number


def games_per_person(games, participants):

    games_per_person_list = [number_of_games(games, p) for p in participants]
    return games_per_person_list


def games_number_is_equal(games, participants):

    if len(set(games_per_person(games, participants))) == 1:
        return True
    else:
        return False


def generate_games():
    tour = get_current_tournament()
    number_of_rounds = get_number_of_rounds(tour)
    divider = 2
    games_number = tour.games_per_person
    participants = [i for i in range(1, tour.number_of_participants() + 1)]

    games = []
    rounds_counter = 0
    attempts_number = 50
    attempt = 0
    games_per_person_list = [0] * len(participants)
    step = 1

    while rounds_counter != number_of_rounds:

        # generate possible couples
        if games_number_is_equal(games, participants):
            games_subset = generate_games_subset(games, participants)
        else:
            min_games = min(games_per_person(games, participants))
            subset_participants = [p for p in participants if number_of_games(games, p) == min_games]

            if len(subset_participants) < divider:
                for p in participants:
                    if p not in subset_participants:
                        subset_participants.append(p)
                        if len(subset_participants) == divider:
                            break

            games_subset = generate_games_subset(games, subset_participants)

        games.extend(games_subset)
        random.shuffle(participants)

        # exit when everybody has the same number of games
        if games_number_is_equal(games, participants):
            rounds_counter += 1

        # start again if impossible to generate unique couples
        attempt += 1
        if attempt > attempts_number:
            participants = [i for i in range(1, tour.number_of_participants() + 1)]
            games = []
            rounds_counter = 0
            attempt = 0

    return games


def generate_playoff_games():
    tournament = get_current_tournament()

    for i in range(4):
        g = Game(tournament=tournament, game_id=i+1, id1=i+1, id2=8-i,
                 game_date=tournament.start_date_playoff, start_time="1{}:00:00".format(i+2))
        g.save()

    semi_day = tournament.start_date_playoff + timedelta(days=1)
    for i in range(2):
        g = Game(tournament=tournament, game_id=i+5, id1=i+1, id2=4-i,
                 game_date=semi_day, start_time="1{}:00:00".format(i*2+3))
        g.save()

    final_day = tournament.start_date_playoff + timedelta(days=2)

    g = Game(tournament=tournament, game_id=7, id1=-5, id2=-6,
             game_date=final_day, start_time="13:00:00")
    g.save()
    g = Game(tournament=tournament, game_id=8, id1=5, id2=6,
             game_date=final_day, start_time="15:00:00")
    g.save()


def get_time(start_time):
    tournament = get_current_tournament()
    TIME_INTERVAL = tournament.game_duration
    # return str(datetime.strptime(start_time, '%H:%M') + timedelta(minutes=TIME_INTERVAL))[11:-3]
    t = datetime.combine(datetime(1,1,1), start_time) + TIME_INTERVAL
    return t.time()


def get_dates():
    tournament = get_current_tournament()
    START_DAY = tournament.start_date
    NUMBER_OF_DAYS = tournament.start_date_playoff - tournament.start_date
    NUMBER_OF_DAYS = NUMBER_OF_DAYS.days
    days = []

    i = 0
    day = START_DAY
    while i < NUMBER_OF_DAYS:

        i += 1
        day_str = day.strftime("%a")

        if "Sun" in day_str or "Sat" in day_str:
            day = day + timedelta(days=1)
            continue
        days.append(day)
        day = day + timedelta(days=1)

    return days


def players_this_day(games, day):
    participants = []
    for game in games:
        if game.game_date == day:
            participants.append(game.id1)
            participants.append(game.id2)

    return participants


def have_slot_for_game(games, game, day, max_games=1):

    players = [game.id1, game.id2]

    for player in players:
        if players_this_day(games, day).count(player) >= max_games:
            return False
    return True


def games_this_day(games_all, day):
    games = []
    for game in games_all:
        if game.game_date == day:
            games.append(game)

    return games


def last_game_time(games_all, day):
    games = games_this_day(games_all, day)
    try:
        games.sort(key=lambda g: datetime.combine(datetime(1,1,1), g.start_time))
    except:
        pass
    return games[-1].start_time


def initial_games_number(days_number):
    t = get_current_tournament()
    nop = t.number_of_participants()
    gpp = t.games_per_person
    if gpp % days_number == 0 and nop % 2 == 0:
        return gpp // days_number
    return gpp // days_number + 1


def generate_schedule(games):
    tournament = get_current_tournament()
    START_TIME = tournament.game_start_time

    days = get_dates()
    num_of_days = len(days)
    not_set_games = []

    for game in games:
        for i in range(num_of_days):
            if have_slot_for_game(games, game, days[i]):
                if len(games_this_day(games, days[i])) == 0:
                    game_time = START_TIME
                else:
                    game_time = get_time(last_game_time(games, days[i]))

                game.start_time = game_time
                game.game_date = days[i]
                break
            if i == num_of_days - 1:
                not_set_games.append(game)

    if len(not_set_games) == 0:
        for game in games:
            game.save()
        return True
    else:
        return False


def split_games_by_days(games):
    games = list(games)
    games.sort(key=lambda game: (game.game_date, game.start_time))
    day = games[0].game_date
    days = []
    days_tmp = []
    for game in games:
        if game.game_date == day:
            days_tmp.append(game)
            continue
        day = game.game_date
        days.append(days_tmp)
        days_tmp = [game]
    days.append(days_tmp)
    return days


def recount_rating():
    tournament = get_current_tournament()
    for p in tournament.participant_set.all():
        p.win_sets = 0
        p.win_balls = 0
        p.games_left = 0
        for g in Game.objects.filter((Q(participant1=p) | Q(participant2=p)) & Q(game_id=0)):
            if g.setresult_set.exists():
                for r in g.setresult_set.all():
                    if g.participant1 == p:
                        if r.result1 > r.result2:
                            p.win_sets += 1
                        p.win_balls += r.result1 - r.result2
                    else:
                        if r.result2 > r.result1:
                            p.win_sets += 1
                        p.win_balls += r.result2 - r.result1
            else:
                p.games_left += 1
        p.save(update_fields=["win_sets", "win_balls", "games_left"])


def write_schedule_to_xls(days, num_of_sets):

    xls_dir = r"tournament\static\tournament"
    filename = "schedule.xlsx"

    book = Workbook()
    sheet = book.active
    sheet.title = "Расписание"

    row = 1
    for day in days:
        fill_flag = True
        for i in range(len(day)):

            # set data
            sheet.cell(row, 1, day[i].game_date)
            sheet.cell(row, 2, day[i].start_time)
            sheet.cell(row, 3, str(day[i].get_p1()))
            sheet.cell(row + 1, 3, str(day[i].get_p2()))

            # set borders
            thin = Side(border_style="thin", color="000000")
            hair = Side(border_style="hair", color="000000")

            top = Border(top=thin)
            left = Border(left=thin)
            right = Border(right=thin)
            bottom = Border(bottom=thin)
            left2 = Border(left=hair)

            end_col = chr(ord('C') + num_of_sets)
            rows = sheet['A' + str(row):end_col + str(row + 1)]

            for cell in rows[0]:
                cell.border = cell.border + top

            for j in range(0, 4):
                rows[0][j].border = rows[0][j].border + left
                rows[1][j].border = rows[1][j].border + left
            for j in range(4, len(rows[0])):
                rows[0][j].border = rows[0][j].border + left2
                rows[1][j].border = rows[1][j].border + left2
            rows[0][-1].border = rows[0][-1].border + right
            rows[1][-1].border = rows[1][-1].border + right

            for cell in rows[-1]:
                cell.border = cell.border + bottom

            # set background
            if fill_flag:
                fill = PatternFill("solid", fgColor="DEDEDE")
                for _row in rows:
                    for cell in _row:
                        cell.fill = fill
                fill_flag = False
            else:
                fill_flag = True

            row += 2
        row += 1
    book.save(os.path.join(BASE_DIR, xls_dir, filename))
